"""Excel ParseAdapter using openpyxl.

Maps each worksheet to one ParsedBlock (text search) and one
ParsedTable (structured cells).  No OCR / layout analysis —
this is the minimal viable implementation.
"""

from __future__ import annotations

import time
from pathlib import Path

from backend.documents.parse_adapter.base import (
    ParsedBlock,
    ParsedDocument,
    ParsedTable,
    ParseError,
    ParseMeta,
)


class ExcelParser:
    """Parse ``.xlsx`` / ``.xls`` files into ParsedDocument."""

    SUPPORTED_EXTENSIONS = {".xlsx", ".xls"}

    def parse(self, file_path: str) -> ParsedDocument:
        path = Path(file_path)
        ext = path.suffix.lower()

        if ext not in self.SUPPORTED_EXTENSIONS:
            raise ParseError(
                f"ExcelParser does not support '{ext}'. "
                f"Supported: {', '.join(sorted(self.SUPPORTED_EXTENSIONS))}"
            )

        if not path.exists():
            raise ParseError(f"File not found: {file_path}")

        try:
            import openpyxl
        except ImportError:
            raise ParseError("openpyxl is not installed. Install with: uv add openpyxl")

        t0 = time.perf_counter()
        warnings: list[str] = []

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ParseError(f"Failed to open Excel file: {exc}") from exc

        blocks: list[ParsedBlock] = []
        tables: list[ParsedTable] = []

        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            rows: list[list[str]] = []
            sheet_text: list[str] = []

            for row in ws.iter_rows(values_only=True):
                row_values = [str(c) if c is not None else "" for c in row]
                if any(v for v in row_values):  # skip fully empty rows
                    rows.append(row_values)
                    line = "\t".join(v for v in row_values if v)
                    sheet_text.append(line)

            if not rows:
                continue

            # Text block for keyword search
            if sheet_text:
                blocks.append(
                    ParsedBlock(
                        block_id=f"excel_sheet_{sheet_index}",
                        page_no=sheet_index,
                        block_type="paragraph",
                        text="\n".join(sheet_text),
                        order_index=sheet_index,
                    )
                )

            # Table with structured cells
            tables.append(
                ParsedTable(
                    table_id=f"excel_table_{sheet_index}",
                    page_no=sheet_index,
                    caption=sheet_name,
                    cells_markdown=_rows_to_markdown(rows),
                    cells_structured=rows,
                )
            )

        wb.close()

        duration_ms = (time.perf_counter() - t0) * 1000
        version = getattr(openpyxl, "__version__", "unknown")

        meta = ParseMeta(
            parse_engine="excel_openpyxl",
            parse_engine_version=version,
            parse_duration_ms=duration_ms,
            total_pages=len(wb.sheetnames),
            parse_warnings=warnings,
            parse_path="native_text",  # M8: Excel is always native text
        )

        return ParsedDocument(
            filename=path.name,
            file_type=ext.lstrip(".").upper(),
            blocks=blocks,
            tables=tables,
            parse_meta=meta,
        )


def _rows_to_markdown(rows: list[list[str]]) -> str:
    """Convert 2-D list to a basic markdown table."""
    if not rows:
        return ""
    lines: list[str] = []
    max_cols = max((len(r) for r in rows), default=0)
    for i, row in enumerate(rows):
        # Pad short rows
        padded = list(row) + [""] * (max_cols - len(row))
        lines.append("| " + " | ".join(padded) + " |")
        if i == 0:
            lines.append("| " + " | ".join("---" for _ in range(max_cols)) + " |")
    return "\n".join(lines)
